import csv
import openpyxl
import json
from enum import Enum
from .models import DataSet
from pprint import pprint

""" Enum to represent filetype, used as an argument for the constructor """
class FileType(Enum):
    CSV=0
    XLSX=1
    JSON=2

""" Exception to throw when given csv/xlsx files have an invalid format"""
class InputException(Exception):
    def __init__(self, message):
        self.message = message

class JSONDataSet:
    """ Class respresented a dataset
    Args:
        filepath (str): Path to a csv/xlsx file on disk (In case of JSON, this is actually the (poorly named) data itself)
        filetype (FileType): Enum value representing the type of the file (CSV/XLSX/JSON)
        name (str): Name for the dataset (can be left empty for FileType.JSON)
        description (str): Description for the dataset (can be left empty for FileType.JSON)
        id (int): The id for this dataset in the database; to be left None if creating a new dataset"""
    def __init__(self, filepath, filetype, name, description, id=None):
        self.name = name # need to save name as an identifier
        self.id = id

        if filetype == FileType.CSV:
            self.json_dict = self._init_csv(filepath, name, description)

        elif filetype == FileType.XLSX:
            self.json_dict = self._init_xlsx(filepath, name, description)

        elif filetype == FileType.JSON:
            # lol why u do this?
            self.json_dict = json.loads(filepath)

    def _init_csv(self, filepath, name, description):
        """Initialize the object with a csv file
        Args:
            filepath (str): Path to the csv file on disk
            name (str): Name of the dataset
            description (str): Description for the dataset"""
        with open(filepath, newline='') as csv_file:
            csv_data = csv.reader(csv_file) # parse the file

            # initialize the data structure
            json_dict = {
                'title': name,
                'description': description,
                'tags': list(),
                'data': {
                    'questions': list(),
                    'responses': list(),
                    'tags': list()}}

            rid = 0 # response id: tracking the current response number

            # for each row in the csv file
            for i, row in enumerate(csv_data):
                if i == 0: # first pass of loop -> first row of csv -> questions
                    for qid, question in enumerate(row):
                        json_dict['data']['questions'].append({
                            'qid': qid,
                            'question': question})

                else: # all rows after the first represent responses to the questions
                    for qid, response in enumerate(row):
                        json_dict['data']['responses'].append({
                            'rid': rid,
                            'qid': qid,
                            'response': response})

                        # need to initialize list of tags for each of the responses,
                        # but that only needs to be done once for each. so check
                        # if this is the first pass through the loop first
                        if qid == 0:
                            json_dict['data']['tags'].append({
                                'rid': rid,
                                'tags': list()})
                    rid += 1

            return json_dict

    def _init_xlsx(self, filepath, name, description):
        """Initialize the object with a xlsx file
        Args:
            filepath (str): Path to the xlsx file on disk
            name (str): Name of the dataset
            description (str): Description for the dataset"""
        xlsx_data = openpyxl.load_workbook(filepath).active

        # initialize the data structure
        json_dict = {
            'title': name,
            'description': description,
            'tags': list(),
            'data': {
                'questions': list(),
                'responses': list(),
                'tags': list(),
            },
        }

        rid = 0 # response id: tracking the current response number

        # for each row in the xlsx file
        for row in range(1, xlsx_data.max_row + 1):
            if row == 1: # first pass of loop -> first row of xlsx -> questions
                for qid, question in enumerate(xlsx_data[str(row)]):
                    json_dict['data']['questions'].append({
                        'qid': qid,
                        'question': question.value})

            else: # all rows after the first represent responses to the questions
                for qid, response in enumerate(xlsx_data[str(row)]):
                    json_dict['data']['responses'].append({
                        'rid': rid,
                        'qid': qid,
                        'response': response.value})
                    # need to initialize list of tags for each of the responses,
                    # but that only needs to be done once for each. so check
                    # if this is the first pass through the loop first
                    if qid == 0:
                        json_dict['data']['tags'].append({
                            'rid': rid,
                            'tags': list()})
                rid += 1
        return json_dict

    # write this dataset to the database
    def SaveDataset(self, user):
        """ Write this dataset to the database; Create a new entry if it isn't already there, and update the old entry if it is
        Args:
            user (User): The user that is saving the dataset"""
        if (self.id == None): # no id defined => this isn't in the database
            dataset = DataSet(data=json.dumps(self.json_dict), user=user)
        else: # this dataset is already in the database, we're just updating it
            dataset = DataSet.objects.get(id=self.id)
            dataset.data = json.dumps(self.json_dict)

        dataset.save()
        self.id = dataset.id

    # get all datasets in the database associated with a specific user
    @staticmethod
    def GetDatasets(user):
        """ Pull all of the datasets from the database corresponding to a specific user
        Args:
            user (User): The user to get datasets for
        Returns:
            List<JSONDataSet>: A list of JSONDataSets"""
        datasets = DataSet.objects.filter(user=user)
        return list(map(lambda x: JSONDataSet(x.data, FileType.JSON, "", "", x.id), datasets))

    @staticmethod
    def GetDataset(id):
        """ Get the dataset with a specific id
        Args:
            id (int): Id of the dataset to get from the database
        Returns:
            JSONDataSet: The dataset with that id from the database"""
        return JSONDataSet(DataSet.objects.get(id=id).data, FileType.JSON, "", "", id)

    def AddTag(self, tag):
        """ Adds a given global tag to the dataset
        Args:
            tag (str): Tag to add to the dataset"""
        self.json_dict["tags"].append(tag)

    def RemoveTag(self, tag):
        """ Removes a given global tag from the dataset, and removes all instances of that tags.
        Does nothing if the tag doesn't exist in the dataset
        Args:
            tag (str): Tag to remove from the dataset"""
        while self.HasTag(tag): # while loop, just in case multiple copies of the tag ended up in the dataset somehow
            self.json_dict["tags"].remove(tag)

        # remove all instances of the tag in the responses
        for response_tags in self.json_dict['data']['tags']:
            while (tag in response_tags['tags']):
                response_tags['tags'].remove(tag)


    def TagItem(self, rid, tags):
        """ Sets the tags of a given response.
        Args:
            rid (int): Response id for the response you are tagging
            tags (list<str>): List of strings representing the tags to set for the response"""
        item = self._GetTagObject(rid)
        item['tags'] = tags
        pprint(item)

    def ItemHasTag(self, rid, tag):
        """ Check if a given response already has a given tag
        Args:
            rid (int): Response id for the response you are checking
            tag (str): Tag you are checking the existence of
        Returns:
            Boolean: True if the tag is already in the response"""
        return tag in self._GetTagObject(rid)['tags']

    def RemoveItemTag(self, rid, tag):
        """ Remove a given tag from a given response
        Args:
            rid (int): Response id for the response you are removing the tag from
            tag (str): Tag you are removing from the response"""
        item = _GetTagObject(rid)
        if (tag in item['tags']):
            item['tags'].remove(tag)

    def HasTag(self, tag):
        """ Check if a given global tag exists
        Args:
            tag (str): Tag you are checking the existence of
        Returns:
            Boolean: True if the tag already exists"""
        return tag in self.json_dict["tags"]

    def SetTags(self,tags):
        """ Set the list of tags on this dataset
        Args:
            tags (list<str>): List of tags to be set to this dataset"""
        self.json_dict['tags'] = tags

    def GetQuestions(self):
        """ Get the list of questions in this DataSet
        Returns:
            list<str>: List of questions"""
        return [question['question'] for question in self.json_dict['data']['questions']]

    def GetTags(self):
        """ Get the list of global tags in the DataSet
        Returns:
            list<str>: List of tags"""
        return self.json_dict['tags']

    def _CheckPermission(self, user):
        """ Check if a requesting user has permission to use this DataSet
        Args:
            user (user): Django user object
        Returns:
            Boolean: True if the user has permission"""
        return DataSet.objects.get(id=self.id).user.id == user.id

    def _GetTagObject(self, rid):
        """ Get the tag object corresponding to a specific response
        Args:
            rid (int): Response id you want the tag object for
        Returns:
            dict(): tag object for the specified response"""
        return next(filter(lambda x: str(x['rid']) == rid, self.json_dict['data']['tags']), dict())

    def GetResponseMatrix(self):
        """ Get a 2D list representation of the responses, that turns out to be useful in certain situations"""
        # initialize the matrix with null values in this silly nested list compehension
        matrix = [[None for i in range(len(self.json_dict['data']['questions']))] for j in range(int(len(self.json_dict['data']['responses'])/len(self.json_dict['data']['questions'])))]
        for response in self.json_dict['data']['responses']:
            # add the values to the matrix
            matrix[response['rid']][response['qid']] = response
            matrix[response['rid']][response['qid']]['tags'] = next(filter(lambda x: x['rid'] == response['rid'], self.json_dict['data']['tags']), dict()).get('tags', list())
        return matrix


    def ExportCSV(self, user):
        """ Export this dataset as a csv file.
        Args:
            user (user): Requesting django user object
        Returns:
            File path to the generated csv file on disk"""
        if not self._CheckPermission(user):
            return

        # create the file in the temp directory with name corresponding to the title
        with open("media/tmp/" + self.json_dict['title'] + ".csv", "w+") as fpntr:
            writer = csv.writer(fpntr) # initialize writer

            # first row: list of questions, then list of tags
            # list comprehension to get the list of questions
            row = [x['question'] for x in self.json_dict['data']['questions']]
            row += self.json_dict['tags'] # concatenate list of questions with list of tags
            # write the row to file
            writer.writerow(row) # write the first row to the csv

            # get the matrix representation of the responses for convenience
            matrix = self.GetResponseMatrix()

            # iterate through the rows
            for row in matrix:
                # construct the {0, 1} representation of tag on responses
                tags_to_append = list()
                for tag in self.json_dict['tags']:
                    tags_to_append.append(1 if tag in row[0]['tags'] else 0)
                # concatenate {0, 1} tags at the end of the reponses
                row = [x['response'] for x in row] + tags_to_append
                writer.writerow(row)

        # return the file path
        return "media/tmp/" + self.json_dict['title'] + ".csv"

    def ExportXLSX(self, user):
        """ Export this dataset as a xlsx file.
        Args:
            user (user): Requesting django user object
        Returns:
            File path to the generated xlsx file on disk"""
        if not self._CheckPermission(user):
            return

        wb = openpyxl.Workbook()
        ws = wb.active

        # first row: list of questions, then list of tags
        # list comprehension to get the list of questions
        row = [x['question'] for x in self.json_dict['data']['questions']]
        row += self.json_dict['tags']  # concatenate list of questions with list of tags
        ws.append(row) # write the first row to the xlsx

        # get the matrix representation of the responses for convenience
        matrix = self.GetResponseMatrix()

        # iterate through the rows
        for row in matrix:
            # construct the {0, 1} representation of tag on responses
            tags_to_append = list()
            for tag in self.json_dict['tags']:
                tags_to_append.append(1 if tag in row[0]['tags'] else 0)
            # concatenate {0, 1} tags at the end of the reponses
            row = [x['response'] for x in row] + tags_to_append
            # write the row to file
            ws.append(row)

        # save and return
        wb.save("media/tmp/" + self.json_dict['title'] + ".xlsx")
        return "media/tmp/self.json_dict['title']" + ".xlsx"

    def DeleteDataSet(self, user):
        if not self._CheckPermission(user):
            return

        dataset = DataSet.objects.get(id=self.id)
        dataset.delete()
